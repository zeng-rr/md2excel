import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class MarkdownToExcelConverter:
    def __init__(self, width: int = 20, height: int = 20, header_font: Font = None, cell_font: Font = None, alignment: Alignment = None, border: Border = None):
        """
        初始化转换器

        :param width: 列宽
        :param height: 行高
        :param header_font: 表头字体样式
        :param cell_font: 单元格字体样式
        :param alignment: 单元格对齐方式
        :param border: 单元格边框
        """
        self.wb = Workbook()
        self.default_sheet = self.wb.active
        self.default_sheet.title = "Sheet1"
        
        # 定义基础样式
        self.width = width
        self.height = height
        self.header_font = Font(bold=True, size=12) if header_font is None else header_font
        self.cell_font = Font(size=11) if cell_font is None else cell_font
        self.alignment = Alignment(horizontal='center', vertical='center') if alignment is None else alignment
        self.border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')) if border is None else border

    def _parse_markdown_tables(self, md_content):
        """解析Markdown中的表格，返回标题和表格数据"""
        # 使用正则表达式匹配markdown表格
        table_pattern = r'(\|.*\|)\n\|([ \-:|]*)\|\n((?:\|.*\|\n)*)'
        tables = []
        
        for i, match in enumerate(re.finditer(table_pattern, md_content)):
            header_row, separator, data_rows = match.groups()
            
            # 获取最近的任意级别标题
            # 获取所有标题及其结束位置
            headers = []
            for h in re.finditer(r'^(#{1,6})\s*(.*?)\s*$', md_content, flags=re.MULTILINE):
                end_pos = h.end()
                if end_pos < match.start():
                    headers.append({
                        'text': h.group(2).strip(),
                        'end_pos': end_pos
                    })
            
            # 找到最近的标题
            closest_header = next(iter(reversed(headers)), None)
            title = closest_header['text'] if closest_header else f'Table {i+1}'
            title = re.sub(r'[^\w\u4e00-\u9fa5]', '', title[:30])
            
            # 处理表头
            header = [cell.strip() for cell in header_row.split('|')[1:-1]]
            
            # 处理数据行
            data = []
            for row in data_rows.split('\n'):
                if row.strip():
                    cells = [cell.strip() for cell in row.split('|')[1:-1]]
                    data.append(cells)
            
            if len(header) > 0 and len(data) > 0:
                tables.append({
                    'title': re.sub(r'[\\/:*?\[\]]', '', title)[:30],
                    'header': header,
                    'data': data
                })
        
        return tables

    def _apply_styles(self, sheet, max_col):
        """应用样式到工作表"""
        # 设置列宽和行高
        for col in range(1, max_col+1):
            column_letter = get_column_letter(col)
            sheet.column_dimensions[column_letter].width = self.width
        
        # 设置行高
        sheet.row_dimensions[1].height = self.height + 5
        for row in range(2, sheet.max_row + 1):
            sheet.row_dimensions[row].height = self.height
        
        # 应用全局样式
        for row in sheet.iter_rows():
            for cell in row:
                cell.font = self.cell_font
                cell.alignment = self.alignment
                cell.border = self.border
        
        # 表头样式
        for cell in sheet[1]:
            cell.font = self.header_font

    def convert(self, md_content, output_path):
        """转换入口函数"""
        tables = self._parse_markdown_tables(md_content)
        
        # 删除默认创建的空sheet
        if len(self.wb.worksheets) > 0 and tables:
            if self.wb.worksheets[0].max_row == 0:  
                self.wb.remove(self.wb.worksheets[0])
        
        # 为每个表格创建sheet
        for i, table in enumerate(tables):
            if i == 0:
                sheet = self.wb.active
                sheet.title = table['title']
            else:
                sheet = self.wb.create_sheet(title=table['title'])
            
            # 写入表头
            sheet.append(table['header'])
            # 写入数据
            for row in table['data']:
                sheet.append(row)
            
            # 应用样式
            self._apply_styles(sheet, len(table['header']))
        
        # 保存文件
        self.wb.save(output_path)

